from openpyxl import load_workbook

wb = load_workbook(".//TestData//Testcase.xlsx")
# ws = wb.active
# p = 'B5'
# # print(ws[p].value)
# c = "boloccc"
# ws[p].value = c
# wb.save(".//TestData//Testcase.xlsx")
# df = pd.read_excel(".//TestData//Testcase.xlsx", )
import pandas as pd

SheetName = "Deb"
Rowno = 0
df = pd.read_excel('.//TestData//Testcase.xlsx', sheet_name=SheetName)
data = df.iloc[Rowno]
# print(df)
print(data)
