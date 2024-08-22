# import xml.etree.cElementTree as ET
#
# mytree = ET.parse('.//TestData//AMZTRN.xml')
# myroot = mytree.getroot()
# for item in myroot.findall("./Shipment/Container/ExternalReference/[IdType='EQUIPMENT_NUMBER']"):
#     c = item.find('Id')
#     b = c.text.encode('utf8')
#     print(b)
#     a = str(b)
#     print(a)
#     d = a.split('b')
#     print(d)
#     w = d[1]
#     print(w)
#     q = w.strip("'")
#     print(q)
#
# for item in myroot.findall("./Shipment/Destination"):
#     v = item.find('Code')
#     h = v.text.encode('utf8')
#     print(h)
#
# for item in myroot.findall("./Shipment/Source"):
#     t = item.find('Code')
#     r = t.text.encode('utf8')
#     print(r)
#     p = h + r
#     print(p)
from openpyxl.reader.excel import load_workbook

wb = load_workbook('C://Users//debadatta//Desktop//New folder//Testing1.xlsm', keep_vba=True)
ws = wb.active
CellName = 'B7'
ws[CellName].value = 3222
wb.save('C://Users//debadatta//Desktop//New folder//Testing1.xlsm')
wb = load_workbook("C://Users//debadatta//Desktop//New folder//Testing1.xlsm")
ws = wb.active
w = ws[CellName].value
print(w)
