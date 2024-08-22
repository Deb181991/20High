import datetime
import json
import time
import xml.etree.cElementTree as ET

import pandas as pd
import pypyodbc
from openpyxl.reader.excel import load_workbook
from selenium.webdriver import ActionChains
from selenium.webdriver.support.wait import WebDriverWait

from Utilities.customLogger import LogGen

myLogger = LogGen.loggeen()


# action = ActionChains(driver)
# actions = ActionChains(browser)

class genericmethod:

    # @ initializing Driver
    def __init__(self, driver):
        self.driver = driver
        self.act = ActionChains(driver)
        # defaultTime = self.driver.implicitly_wait(60)
        # self.driver.implicitly_wait(60)

    # @ send keys to text field
    # @ element
    def type(self, element, text):
        element.send_keys(text)
        myLogger.info("******Enter Data ******")

    # @ send keys to text field
    # @ element
    def click(self, element):
        element.click()
        myLogger.info("******Clicking Element ******")

    # @ Click using Javascript
    # @ element
    def clickUsingJS(self, element):
        myLogger.info("clicking using JS")
        self.driver.execute_script('arguments[0].click()', element)
        # self.waitForPageToLoad()
        # self.waitForElementToBePresent(element,elementName)
        # myLogger.info("Clicking at " + elementName + "")
        # self.waitForPageToLoad()

    # @ send keys to text field
    # @ element
    def wait(self, TimeUnit):
        time.sleep(TimeUnit)
        myLogger.info("****** Waitinging ******")

    # @ send keys to text field
    # @ element
    def navigateTo(self, url):
        myLogger.info("Navigating to " + url + "")
        self.driver.get(url)

    # @ get Text from element
    # @ element
    def getText(self, element):
        myLogger.info("************ Retrieving text *************")
        data = element.text
        # driver.log(data)
        myLogger.info(data)
        return data

    # @ get Value from text field
    # @ element
    def getValue(self, element):
        myLogger.info("************ Retrieving text *************")
        data = element.get_property('value')
        # driver.log(data)
        myLogger.info(data)
        return data

    # @ Waiting for element to be visible
    # @ element
    def waitForElementToBeVisible(self, element, elementname):
        myLogger.info("Waiting for " + elementname + " to be visible")
        myLogger.info("************* Waiting For Element *************")
        element.is_displayed  # (self.driver.implicitly_wait(60))
        assert True

    # @ Waiting for element to be Present
    # @ element
    def waitForElementToBePresent(self, element, elementname):
        myLogger.info("Waiting for " + elementname + " to be present")
        element.is_displayed(self.driver.implicitly_wait(60))
        assert True

    # @ Waiting for element to be click
    # @ element
    def waitForElementToBeClickable(self, element, elementname):
        myLogger.info("Waiting for " + elementname + " to be clickable")
        element.is_displayed
        assert True

    # @ Waiting for page to be Load
    # @ element
    def waitForPageToLoad(self):
        myLogger.info("waiting for page to load")
        WebDriverWait(self.driver, self.SERVER_TIMEOUT).until(
            lambda wd: self.driver.execute_script("return document.readyState") == 'complete',
            "Page taking too long to load"
        )

    # @ get data from file
    def ReadData(self, filePath, object):
        file = open(filePath, "r")
        # x = file.read()
        finaldata = json.load(file)
        data = finaldata[object]
        print(finaldata)
        print(data)
        return data

    # @ Write data in a file
    def WriteData(self, filePath, object, value):
        data = {}
        data[object] = value
        j = json.dumps(data)
        with open(filePath, "w") as f:
            f.write(j)
        return value

    # @ Perform mouse move action and click the target element
    # @ element
    def MouseHoverAndClick(self, element):
        self.act.move_to_element(element).click().perform()
        myLogger.info("****** MouseHover Click an Element ******")

    # @ Perform mouse move action and click the target element
    # @ element
    def MouseHover(self, element):
        self.act.move_to_element(element).perform()
        myLogger.info("****** MouseHover Click an Element ******")

    def ScrollDown(self):
        self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
        myLogger.info("****** scrolled down ******")

    def Increment(self, element):
        file = open(".//TestData//data.json", "r")
        finaldata = json.load(file)
        print(finaldata)
        print(finaldata['Numeric'])
        a = finaldata['Numeric']
        print(a)
        b = int(a) + 1
        print(b)
        data = {}
        data['Numeric'] = b
        j = json.dumps(data)
        with open(".//TestData//data.json", "w") as f:
            f.write(j)
        element.send_keys(b)

    def NumIncrement(self):
        file = open(".//TestData//data.json", "r")
        finaldata = json.load(file)
        print(finaldata)
        print(finaldata['Numeric'])
        a = finaldata['Numeric']
        print(a)
        b = int(a) + 1
        print(b)
        data = {}
        data['Numeric'] = b
        j = json.dumps(data)
        with open(".//TestData//data.json", "w") as f:
            f.write(j)
        return b

    def switchToTab(self, windowCount):
        self.driver.switch_to.window(self.driver.window_handles[windowCount])

    def Tab(self):
        self.driver.execute_script('window.open()')

    def current_url(self):
        var = self.driver.current_url
        myLogger.info(var)
        self.driver.get(var)

    def duplicate_Tab(self, windowCount):
        var = self.driver.current_url
        self.driver.execute_script('window.open()')
        self.driver.switch_to.window(self.driver.window_handles[windowCount])
        self.driver.get(var)

    def XML_READ(self):
        mytree = ET.parse('.//TestData//AMZTRN.xml')
        myroot = mytree.getroot()
        for item in myroot.findall("./Shipment/Container/ExternalReference/[IdType='EQUIPMENT_NUMBER']"):
            c = item.find('Id')
            b = c.text.encode('utf8')
            print(b)
            a = str(b)
            print(a)
            d = a.split('b')
            print(d)
            w = d[1]
            q = w.strip("'")
            print(q)
            myLogger.info(q)
            return q

    def sqldata(self):
        mytree = ET.parse('.//TestData//AMZTRN.xml')
        myroot = mytree.getroot()

        for item in myroot.findall("./Shipment/Destination"):
            v = item.find('Code')
            h = v.text.encode('utf8')

        for item in myroot.findall("./Shipment/Source"):
            t = item.find('Code')
            r = t.text.encode('utf8')
            p = h + r
            v = str(p)
            w = v.strip("b")

        con = pypyodbc.connect("Driver={SQL Server};"
                               "Server=devASQLWEBEC01;"
                               "Database=ADM_Archive;"
                               "uid=web_merge_ec;pwd=w3bm3rg33c")

        query = "SELECT iOriginalDest_ID FROM webfts_hudd..tblFTS_EDI_PortOfDischarge WITH (NOLOCK)where vcPortofDischarge like" + " " + w
        df = pd.read_sql_query(query, con)
        df.to_excel('.//TestData//Testcase.xlsx')
        h = pd.read_excel('.//TestData//Testcase.xlsx', sheet_name='Sheet1')
        c = h['ioriginaldest_id'].unique()
        ta = str(c)
        ta1 = ta.strip("[").strip("]")
        query1 = "select Dest_Name from WEBEC.dbo.comm_Destination_tbl where Dest_ID like" + " " + ta1
        df1 = pd.read_sql_query(query1, con)
        df1.to_excel('.//TestData//Testcase.xlsx', sheet_name='Sheet2')
        df2 = pd.read_excel('.//TestData//Testcase.xlsx', sheet_name='Sheet2')
        df3 = df2['dest_name'].unique()
        df4 = str(df3)
        df5 = df4.strip("[").strip("]").strip("'").strip("'")
        myLogger.info("********DF5**********" + df5)
        return df5

    def Refresh(self):
        self.driver.refresh()

    def Clear(self, element):
        element.clear()

    def CurrentDay(self):
        current_time = datetime.datetime.now()
        today = current_time.day
        b = current_time.year
        # today = date.today()
        myLogger.info(today)
        # print(today)

    def Upload(self, element, Path):
        element.send_keys(Path)

    def defaultTime(self):
        self.driver.implicitly_wait(65)

    def READEXCELDATA(self, CellName):
        wb = load_workbook(".//TestData//Testcase.xlsx")
        ws = wb.active
        return ws[CellName].value

    def WRITEEXCELDATA(self, CellName, Value):
        wb = load_workbook(".//TestData//Testcase.xlsx")
        # wb = load_workbook('E://Phoenix_20High//Upload//XBEES.xlsm', keep_vba=True)
        ws = wb.active
        ws[CellName].value = Value
        return wb.save(".//TestData//Testcase.xlsx")
        # return wb.save("E://Phoenix_20High//Upload//XBEES.xlsm")

    def READEXCELDATA_XLSM(self, CellName):
        wb = load_workbook('E://Phoenix_20High//Upload//XBEES.xlsm', keep_vba=True)
        ws = wb.active
        return ws[CellName].value

    def WRITEEXCELDATA_XLSM(self, CellName, Value):
        wb = load_workbook('E://Phoenix_20High//Upload//XBEES.xlsm', keep_vba=True)
        ws = wb.active
        ws[CellName].value = Value
        return wb.save("E://Phoenix_20High//Upload//XBEES.xlsm")

    # def SCREEN_Record(self):
    #     height = GetSystemMetrics(1)
    #     width = GetSystemMetrics(0)
    #     time_stamp = datetime.datetime.now().strftime(
    #         '%Y-%m-%d %H-%M-%S')
    #     file_name = f'{time_stamp}.mp4'
    #     fourcc = cv2.VideoWriter_fourcc('m', 'p', '4', 'v')
    #     final_video = cv2.VideoWriter(file_name, fourcc, 20.0, (width, height))
    #     while True:
    #         img = ImageGrab.grab(bbox=(
    #             0, 0, width, height))
    #         img_ny = ny.array(img)
    #         img_final = cv2.cvtColor(img_ny, cv2.COLOR_BGR2RGB)
    #         final_video.write(img_final)
    #         if cv2.waitKey(10) == ord('t'):
    #             break
